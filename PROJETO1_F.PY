import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import os
from tkinter import messagebox

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

root = ctk.CTk()
root.title("Formulário de Cadastro")
root.geometry('500x400+300+200')
root.resizable(False, False)


def salvar_dados():
    nome = entrada_nome.get()
    idade = entrada_idade.get()
    email = entrada_email.get()
    cpf = entrada_cpf.get()

    if nome and idade and email and cpf:
        if os.path.exists("dados.xlsx"):
            workbook = load_workbook("dados.xlsx")
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Nome", "Idade", "E-mail", "CPF"])

        sheet.append([nome, idade, email, cpf])
        workbook.save("dados.xlsx")

        entrada_nome.delete(0, "end")
        entrada_idade.delete(0, "end")
        entrada_email.delete(0, "end")
        entrada_cpf.delete(0, "end")

        messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")
    else:
        messagebox.showerror("Erro", "Preencha todos os campos!")


frame = ctk.CTkFrame(root, corner_radius=15)
frame.pack(pady=20, padx=20, fill="both", expand=True)

titulo = ctk.CTkLabel(
    frame, text="Formulário de Cadastro", font=("Arial", 24, "bold"))
titulo.pack(pady=20, padx=10)

entrada_nome = ctk.CTkEntry(
    frame, placeholder_text="Nome", width=300, height=40, corner_radius=10
)
entrada_nome.pack(pady=10, padx=10)

entrada_idade = ctk.CTkEntry(
    frame, placeholder_text="Idade", width=300, height=40, corner_radius=10
)
entrada_idade.pack(pady=10, padx=10)

entrada_email = ctk.CTkEntry(
    frame, placeholder_text="E-mail", width=300, height=40, corner_radius=10
)
entrada_email.pack(pady=10, padx=10)

entrada_cpf = ctk.CTkEntry(
    frame, placeholder_text="CPF", width=300, height=40, corner_radius=10
)
entrada_cpf.pack(pady=10, padx=10)

botao_salvar = ctk.CTkButton(
    frame, text="Salvar", command=salvar_dados, width=300, height=40
)
botao_salvar.pack(pady=10, padx=10)

root.mainloop()
